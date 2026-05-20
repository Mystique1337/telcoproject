"""LLM-as-Judge over the 50-pair human-eval set.

Replaces / supplements human raters with LLM judges using the same blind
A/B pair format. Standard methodology from MT-Bench (Zheng et al. 2023):

  - For each pair, ask each judge: "Which review is more authentic?"
  - Run each pair TWICE per judge with sides swapped (position-bias control)
  - Aggregate to per-judge win-rate + Wilson 95% CI
  - Compute Fleiss kappa for inter-judge agreement

Judges (defaults):
  - openai:gpt-5.5
  - anthropic:claude-sonnet-4-20250514
  - hf:meta-llama/Llama-3.3-70B-Instruct

Inputs:
  paper/human_eval_template.xlsx    — same blind file humans would fill
  paper/human_eval_pairs_master.json — answer key (which model on each side)

Outputs:
  paper/llm_judge_summary.json
  paper/llm_judge_summary.md

Run:
  python scripts/llm_judge.py                    # 3 judges, all 50 pairs
  python scripts/llm_judge.py --n 10             # quick smoke
  python scripts/llm_judge.py --judges gpt5,claude  # subset
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import math
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Manual .env load to avoid frame-based path discovery issues
for line in (PROJECT_ROOT / ".env").read_text().splitlines():
    line = line.strip()
    if not line or line.startswith("#") or "=" not in line:
        continue
    k, _, v = line.partition("=")
    os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

import openpyxl

from app.llm.client import get_llm_client, LLMError

logger = logging.getLogger("llm_judge")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s — %(message)s")

XLSX_PATH    = PROJECT_ROOT / "paper" / "human_eval_template.xlsx"
MASTER_PATH  = PROJECT_ROOT / "paper" / "human_eval_pairs_master.json"
OUT_JSON     = PROJECT_ROOT / "paper" / "llm_judge_summary.json"
OUT_MD       = PROJECT_ROOT / "paper" / "llm_judge_summary.md"

JUDGES = {
    "gpt5_5":    "openai:gpt-5.5",
    "claude":    "anthropic:claude-sonnet-4-20250514",
    "llama_70b": "hf:meta-llama/Llama-3.3-70B-Instruct",
}

MODEL_A_NAME = "lmstudio:naija-reviewer-8b"   # the model we want to score wins for


# --------------------------------------------------------------------------- #
# Data loading                                                                  #
# --------------------------------------------------------------------------- #

def load_pairs() -> list[dict[str, Any]]:
    """Merge xlsx (review text) + master.json (answer key) into one pair list."""
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"missing {XLSX_PATH} — run build_human_eval_xlsx.py first")
    if not MASTER_PATH.exists():
        raise FileNotFoundError(f"missing {MASTER_PATH}")
    master = {m["pair_id"]: m for m in json.loads(MASTER_PATH.read_text())}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Evaluation"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    pairs: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[idx["pair_id"]]
        if not pid or pid not in master:
            continue
        pairs.append({
            "pair_id":          pid,
            "persona_id":       row[idx["persona_id"]],
            "register_tier":    row[idx["register_tier"]],
            "location":         row[idx["location"]],
            "product_title":    row[idx["product_title"]],
            "product_category": row[idx["product_category"]],
            "review_a_text":    row[idx["Review A"]] or "",
            "review_a_rating":  row[idx["Rating A"]],
            "review_b_text":    row[idx["Review B"]] or "",
            "review_b_rating":  row[idx["Rating B"]],
            "left_model":       master[pid]["left_model"],
            "right_model":      master[pid]["right_model"],
        })
    return pairs


# --------------------------------------------------------------------------- #
# Judge prompt                                                                  #
# --------------------------------------------------------------------------- #

JUDGE_SYSTEM = (
    "You are an expert evaluator of Nigerian product reviews. You have deep "
    "fluency in the four Nigerian register tiers (Pidgin / code-mixed / "
    "Nigerian English / standard English) and the cultural intensity "
    "calibration that Nigerian users apply when rating products. Your job is "
    "to compare two anonymous reviews and decide which one a real Nigerian "
    "of the described persona is more likely to have written. You output "
    "exactly one token: A, B, or Equal. No prose, no JSON, no markdown."
)

JUDGE_PROMPT_TEMPLATE = """You are evaluating two anonymous product reviews of the same item, both
purportedly written by the same Nigerian user. Pick the one that more
authentically reflects this user.

PERSONA
  persona_id:      {persona_id}
  register_tier:   {register_tier}
  location:        {location}

PRODUCT
  title:           {product_title}
  category:        {product_category}

REVIEW A
  rating:  {rating_a}/5
  text:    "{text_a}"

REVIEW B
  rating:  {rating_b}/5
  text:    "{text_b}"

CRITERIA (in priority order):
  1. Register fidelity — does the review match the user's stated register
     tier and use markers a real {register_tier} speaker would use?
  2. Persona consistency — does the rating + content match what someone in
     this location / occupation / cultural context would actually write?
  3. Authenticity — does it sound like a Nigerian person writing on Jumia,
     not a sanitised English template?
  4. Cultural specificity — Owambe / Pidgin / religious markers /
     Naija-specific concerns appear where appropriate.

Reply with EXACTLY ONE WORD: A, B, or Equal. No other text."""


def _parse_verdict(text: str) -> str:
    """Tolerant parse: pick first occurrence of A / B / Equal token."""
    if not text:
        return "Equal"
    t = text.strip().upper()
    # First non-whitespace alphanumeric character or 'EQUAL' phrase
    import re as _re
    m = _re.search(r"\b(A|B|EQUAL)\b", t)
    if m:
        v = m.group(1)
        return v if v in ("A", "B") else "Equal"
    # Single-letter fallback
    if t.startswith("A"):
        return "A"
    if t.startswith("B"):
        return "B"
    return "Equal"


async def _judge_one(client, system: str, prompt: str, max_tries: int = 2,
                       max_tokens: int = 2000) -> tuple[str, str]:
    """Single judge call with retry.

    Returns (verdict, raw_response). Reasoning models (GPT-5/5.5/o1/o3) burn
    most of their token budget on hidden reasoning, so we use a generous
    `max_tokens` default. The parser only looks for the first A/B/EQUAL token.
    """
    last_err: Exception | None = None
    for attempt in range(max_tries):
        try:
            raw = await client.complete(
                prompt, system=system, max_tokens=max_tokens, temperature=0.0,
            )
            return _parse_verdict(raw), (raw or "")
        except Exception as e:  # noqa: BLE001
            last_err = e
            await asyncio.sleep(1.5 ** attempt)
    logger.warning("judge call failed after %d tries: %s", max_tries, last_err)
    return "Equal", f"<error: {last_err}>"


# --------------------------------------------------------------------------- #
# Evaluation                                                                    #
# --------------------------------------------------------------------------- #

async def evaluate_pair(judge_client, pair: dict, sides_flipped: bool) -> dict:
    """One judge call on one pair, optionally with sides swapped for
    position-bias control."""
    if sides_flipped:
        text_a, text_b = pair["review_b_text"], pair["review_a_text"]
        rating_a, rating_b = pair["review_b_rating"], pair["review_a_rating"]
    else:
        text_a, text_b = pair["review_a_text"], pair["review_b_text"]
        rating_a, rating_b = pair["review_a_rating"], pair["review_b_rating"]

    prompt = JUDGE_PROMPT_TEMPLATE.format(
        persona_id=pair["persona_id"],
        register_tier=pair["register_tier"],
        location=pair["location"] or "",
        product_title=pair["product_title"],
        product_category=pair["product_category"] or "",
        rating_a=rating_a, text_a=text_a[:1500],
        rating_b=rating_b, text_b=text_b[:1500],
    )
    verdict, raw = await _judge_one(judge_client, JUDGE_SYSTEM, prompt)

    # Map verdict back to the actual model (account for the flip).
    # In flipped view: 'A' in the prompt = original right side
    if verdict == "A":
        picked = pair["right_model"] if sides_flipped else pair["left_model"]
    elif verdict == "B":
        picked = pair["left_model"]  if sides_flipped else pair["right_model"]
    else:
        picked = "Equal"

    return {
        "pair_id": pair["pair_id"],
        "sides_flipped": sides_flipped,
        "verdict": verdict,
        "picked_model": picked,
        "raw_response": raw[:500],
    }


async def run_judge(judge_name: str, judge_spec: str, pairs: list[dict],
                     concurrency: int) -> list[dict]:
    """Run one judge over all pairs × 2 orientations."""
    logger.info("→ judge %s (%s) on %d pairs × 2 orientations", judge_name, judge_spec, len(pairs))
    client = get_llm_client(judge_spec)
    sem = asyncio.Semaphore(concurrency)

    async def _bound(pair, flipped):
        async with sem:
            return await evaluate_pair(client, pair, flipped)

    tasks = []
    for p in pairs:
        tasks.append(_bound(p, False))
        tasks.append(_bound(p, True))
    return await asyncio.gather(*tasks)


# --------------------------------------------------------------------------- #
# Stats                                                                         #
# --------------------------------------------------------------------------- #

def wilson_ci(wins: int, n: int, confidence: float = 0.95) -> tuple[float, float]:
    if n == 0:
        return (float("nan"), float("nan"))
    p = wins / n
    z = 1.96
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n) / denom
    return (max(0.0, centre - half), min(1.0, centre + half))


def fleiss_kappa(judge_verdicts: dict[str, list[dict]]) -> float:
    """Fleiss κ over per-pair categorical agreement, treating each (pair_id,
    orientation) as one item and the judge votes as the raters."""
    # Build {(pair_id, flipped): [verdict_judge1, verdict_judge2, ...]}
    by_item: dict[tuple, list[str]] = defaultdict(list)
    for judge, results in judge_verdicts.items():
        for r in results:
            by_item[(r["pair_id"], r["sides_flipped"])].append(r["verdict"])

    items = list(by_item.values())
    items = [v for v in items if len(v) == len(judge_verdicts)]  # only complete rows
    if not items:
        return float("nan")
    n = len(items)
    n_raters = len(judge_verdicts)
    categories = sorted({v for vs in items for v in vs})
    if len(categories) < 2 or n_raters < 2:
        return float("nan")

    # P_i = (1/(n*(n-1))) * (sum n_ij^2 - n) per item
    P_i = []
    for vs in items:
        c = Counter(vs)
        s = sum(v * v for v in c.values())
        P_i.append((s - n_raters) / (n_raters * (n_raters - 1)))
    P_bar = sum(P_i) / n

    # P_e = sum (p_j)^2 where p_j = fraction of total ratings in category j
    total_ratings = n * n_raters
    cat_counts = Counter(v for vs in items for v in vs)
    P_e = sum((cat_counts[c] / total_ratings) ** 2 for c in categories)
    if P_e == 1.0:
        return float("nan")
    return (P_bar - P_e) / (1 - P_e)


def aggregate(judge_results: dict[str, list[dict]]) -> dict:
    """Compute per-judge + overall win rates + Wilson CIs + Fleiss kappa."""
    summary: dict[str, Any] = {
        "judges": list(judge_results.keys()),
        "n_pairs": len(set(r["pair_id"] for v in judge_results.values() for r in v)),
        "model_a": MODEL_A_NAME,
        "per_judge": {},
    }

    for judge, results in judge_results.items():
        naija_wins, claude_wins, ties = 0, 0, 0
        for r in results:
            pm = r["picked_model"]
            if pm == MODEL_A_NAME:
                naija_wins += 1
            elif pm == "Equal":
                ties += 1
            else:
                claude_wins += 1
        decisive = naija_wins + claude_wins
        wlo, whi = wilson_ci(naija_wins, decisive)
        summary["per_judge"][judge] = {
            "naija_wins":    naija_wins,
            "claude_wins":   claude_wins,
            "ties":          ties,
            "decisive":      decisive,
            "naija_win_rate": (naija_wins / decisive) if decisive else None,
            "naija_win_rate_ci95": [wlo, whi],
        }

    # Majority vote across judges per (pair, orientation)
    by_item: dict[tuple, list[str]] = defaultdict(list)
    for judge, results in judge_results.items():
        for r in results:
            by_item[(r["pair_id"], r["sides_flipped"])].append(r["picked_model"])

    naija_maj, claude_maj, tie_maj = 0, 0, 0
    for picks in by_item.values():
        c = Counter(picks)
        top, top_n = c.most_common(1)[0]
        if top_n < (len(picks) + 1) / 2:
            tie_maj += 1
            continue
        if top == MODEL_A_NAME:
            naija_maj += 1
        elif top == "Equal":
            tie_maj += 1
        else:
            claude_maj += 1
    decisive_maj = naija_maj + claude_maj
    wlo, whi = wilson_ci(naija_maj, decisive_maj)
    summary["overall_majority"] = {
        "naija_wins":         naija_maj,
        "claude_wins":        claude_maj,
        "ties":               tie_maj,
        "decisive":           decisive_maj,
        "naija_win_rate":     (naija_maj / decisive_maj) if decisive_maj else None,
        "naija_win_rate_ci95": [wlo, whi],
    }

    summary["fleiss_kappa"] = fleiss_kappa(judge_results)
    return summary


def write_outputs(judge_results: dict[str, list[dict]], summary: dict) -> None:
    OUT_JSON.write_text(json.dumps({
        "summary": summary,
        "raw_results": judge_results,
    }, indent=2, default=str))
    logger.info("💾 %s", OUT_JSON)

    md = ["# LLM-as-Judge: NaijaReviewer-8B vs Claude Sonnet 4", "",
          "_Same 50 blind A/B pairs used for human evaluation. Each judge votes "
          "on each pair TWICE (sides swapped) for position-bias control._", "",
          f"- **Pairs evaluated**: {summary['n_pairs']}",
          f"- **Judges**: {', '.join(summary['judges'])}",
          f"- **Inter-judge Fleiss κ** (nominal): **{summary['fleiss_kappa']:.3f}**"
          if summary['fleiss_kappa'] == summary['fleiss_kappa'] else "- Fleiss κ: n/a",
          "",
          "## Per-judge breakdown",
          "",
          "| Judge | NaijaReviewer wins | Claude wins | Ties | Decisive | NaijaReviewer win-rate | 95% CI |",
          "|---|---|---|---|---|---|---|",
    ]
    for j, r in summary["per_judge"].items():
        wlo, whi = r["naija_win_rate_ci95"]
        wr = r["naija_win_rate"]
        md.append(
            f"| {j} | {r['naija_wins']} | {r['claude_wins']} | {r['ties']} | "
            f"{r['decisive']} | "
            + (f"**{wr:.1%}**" if wr is not None else "—")
            + f" | [{wlo:.1%}, {whi:.1%}] |"
        )

    o = summary["overall_majority"]
    wlo, whi = o["naija_win_rate_ci95"]
    md += [
        "",
        "## Majority vote across all judges (per pair × orientation)",
        "",
        f"- NaijaReviewer wins: **{o['naija_wins']}**",
        f"- Claude wins:        **{o['claude_wins']}**",
        f"- Ties / no majority: **{o['ties']}**",
        (f"- NaijaReviewer **win-rate {o['naija_win_rate']:.1%} (95% CI "
         f"[{wlo:.1%}, {whi:.1%}])**"
         if o['naija_win_rate'] is not None else ""),
        "",
        "## Interpretation",
        "",
        "Following Zheng et al. 2023 (MT-Bench / LLM-as-Judge methodology), "
        "we run each pair twice with sides swapped to control for position "
        "bias, and aggregate across multiple independent judges to control "
        "for any single judge's idiosyncrasies. The Fleiss κ above measures "
        "inter-judge agreement (>0.4 = moderate agreement; >0.6 = substantial; "
        ">0.8 = near-perfect).",
        "",
        "**This methodology supplements but does not replace human evaluation. "
        "Human rater XLSX template at `paper/human_eval_template.xlsx` is still "
        "the ground-truth measurement; LLM-judge is reported as a high-throughput "
        "screening pass for behavioural fidelity.**",
    ]
    OUT_MD.write_text("\n".join(md))
    logger.info("💾 %s", OUT_MD)


# --------------------------------------------------------------------------- #
# Main                                                                          #
# --------------------------------------------------------------------------- #

async def main_async(args) -> int:
    pairs = load_pairs()
    if args.n:
        pairs = pairs[:args.n]
    logger.info("loaded %d pairs from xlsx + master", len(pairs))

    judges = JUDGES
    if args.judges:
        wanted = set(args.judges.split(","))
        judges = {k: v for k, v in JUDGES.items() if k in wanted}
        if not judges:
            logger.error("no judges matched --judges %s; available: %s",
                         args.judges, list(JUDGES))
            return 1

    judge_results: dict[str, list[dict]] = {}
    for jname, jspec in judges.items():
        try:
            judge_results[jname] = await run_judge(jname, jspec, pairs, args.concurrency)
        except Exception as e:  # noqa: BLE001
            logger.error("judge %s failed entirely: %s", jname, e)

    if not judge_results:
        logger.error("no judges completed")
        return 1

    summary = aggregate(judge_results)
    write_outputs(judge_results, summary)

    print()
    print(OUT_MD.read_text())
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--n", type=int, default=None,
                    help="Optional: cap pairs (default = all 50)")
    ap.add_argument("--judges", type=str, default=None,
                    help="Comma list of judge keys (default = all)")
    ap.add_argument("--concurrency", type=int, default=4)
    args = ap.parse_args()
    return asyncio.run(main_async(args))


if __name__ == "__main__":
    sys.exit(main())
