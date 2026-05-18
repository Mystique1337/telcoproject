"""Generate a shareable Excel workbook for human eval of NaijaReviewer-8B
vs Claude Sonnet 4 (or any two backbones).

Output:
  paper/human_eval_template.xlsx       — shareable, NO model labels visible
  paper/human_eval_pairs_master.json   — local-only, has model→side mapping

Workbook structure (3 sheets):
  1. Instructions — what to do, how to vote, time estimate.
  2. Personas — quick reference table of the 24 persona archetypes.
  3. Evaluation — N rows, each with persona/product context, review A,
     review B, and a 'Your vote' dropdown.

Vote values: A_better / B_better / Equal / Skip.

Pairs are randomly side-swapped per row (50/50 chance A is the fine-tune
or Claude) so raters can't pattern-match.

Run:
  # Default: 50 pairs across all 24 personas
  python scripts/build_human_eval_xlsx.py

  # More samples for stronger statistics
  python scripts/build_human_eval_xlsx.py --n 75

Then share `paper/human_eval_template.xlsx` with teammates; each fills in
the 'Your vote' column and (optionally) the 'Comments' column. Renamed
copies (e.g.\ `human_eval_ashinze.xlsx`) come back to the repo's
paper/human_eval_returned/ directory, and `aggregate_human_eval_xlsx.py`
tallies them.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import random
import sys
from pathlib import Path
from typing import Any

import httpx

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("xlsx_eval")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s — %(message)s")

API_URL = "http://localhost:8765"
PERSONAS_DIR = PROJECT_ROOT / "data" / "sample" / "personas"
PRODUCTS_DIR = PROJECT_ROOT / "data" / "sample" / "products"

OUT_XLSX = PROJECT_ROOT / "paper" / "human_eval_template.xlsx"
OUT_MASTER = PROJECT_ROOT / "paper" / "human_eval_pairs_master.json"

MODEL_A_DEFAULT = "lmstudio:naija-reviewer-8b"
MODEL_B_DEFAULT = "anthropic:claude-sonnet-4-20250514"


# --------------------------------------------------------------------------- #
# Data loading                                                                 #
# --------------------------------------------------------------------------- #

def _load_personas() -> list[dict]:
    out = []
    for p in sorted(PERSONAS_DIR.glob("*.json")):
        try:
            out.append(json.loads(p.read_text(encoding="utf-8")))
        except Exception:
            continue
    return out


def _load_products(rng: random.Random, n_sample: int) -> list[dict]:
    files = list(PRODUCTS_DIR.glob("*.json"))
    rng.shuffle(files)
    out = []
    for p in files[: n_sample * 3]:
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            if d.get("title") and d.get("description") and len(d["description"]) > 60:
                out.append(d)
        except Exception:
            continue
        if len(out) >= n_sample:
            break
    return out


# --------------------------------------------------------------------------- #
# API calls (both backbones in parallel per pair)                              #
# --------------------------------------------------------------------------- #

async def _gen(client: httpx.AsyncClient, persona: dict, product: dict,
                backbone: str) -> dict[str, Any]:
    try:
        r = await client.post(
            f"{API_URL}/simulate-review",
            json={"persona": persona, "product": product, "backbone_override": backbone},
            timeout=180,
        )
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e)[:200]}


async def _build_pair(client: httpx.AsyncClient, sem: asyncio.Semaphore,
                       persona: dict, product: dict,
                       model_a: str, model_b: str) -> dict | None:
    async with sem:
        a_task = _gen(client, persona, product, model_a)
        b_task = _gen(client, persona, product, model_b)
        a, b = await asyncio.gather(a_task, b_task)
    if "error" in a or "error" in b:
        logger.warning("  pair failed: %s × %s → A=%s B=%s",
                       persona.get("user_id", "?"), product.get("product_id", "?")[:30],
                       "ok" if "error" not in a else "err",
                       "ok" if "error" not in b else "err")
        return None
    return {"persona": persona, "product": product,
            "model_a_review": a, "model_b_review": b}


async def build_all_pairs(n_pairs: int, model_a: str, model_b: str,
                            concurrency: int, seed: int) -> list[dict]:
    rng = random.Random(seed)
    personas = _load_personas()
    products = _load_products(rng, n_pairs * 3)
    if not personas or not products:
        logger.error("missing personas or products under data/sample/")
        return []

    # Stratified sample: cycle through personas so each gets ~equal coverage
    persona_cycle = list(personas)
    rng.shuffle(persona_cycle)
    combos: list[tuple[dict, dict]] = []
    for i in range(n_pairs):
        p = persona_cycle[i % len(persona_cycle)]
        prod = rng.choice(products)
        combos.append((p, prod))

    logger.info("generating %d pairs   A=%s   B=%s   concurrency=%d",
                n_pairs, model_a, model_b, concurrency)
    sem = asyncio.Semaphore(concurrency)
    async with httpx.AsyncClient() as client:
        results = await asyncio.gather(*[
            _build_pair(client, sem, p, pr, model_a, model_b)
            for p, pr in combos
        ])

    pairs: list[dict] = []
    for i, r in enumerate(results):
        if r is None:
            continue
        # Side-swap with 50% probability so blind raters can't pattern-match
        swap = rng.random() < 0.5
        left_review  = r["model_b_review"] if swap else r["model_a_review"]
        right_review = r["model_a_review"] if swap else r["model_b_review"]
        left_model   = model_b if swap else model_a
        right_model  = model_a if swap else model_b
        pairs.append({
            "pair_id": f"p{len(pairs)+1:03d}",
            "persona_id": r["persona"].get("user_id"),
            "register_tier": r["persona"].get("register_tier"),
            "demographics": r["persona"].get("demographics", {}),
            "product_title": r["product"].get("title", "")[:120],
            "product_category": r["product"].get("category", ""),
            "product_price": r["product"].get("price_naira"),
            "left":  {"review": left_review.get("review", ""),
                       "rating": left_review.get("rating", 3),
                       "_model": left_model},
            "right": {"review": right_review.get("review", ""),
                       "rating": right_review.get("rating", 3),
                       "_model": right_model},
        })
    logger.info("✅ %d pairs built (of %d requested)", len(pairs), n_pairs)
    return pairs


# --------------------------------------------------------------------------- #
# Excel writer                                                                 #
# --------------------------------------------------------------------------- #

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _write_instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 100

    lines = [
        ("Naija Persona Agent — Human Evaluation", True, 18),
        ("", False, 11),
        ("What is this?", True, 14),
        (
            "We are evaluating two LLM-based agents that simulate Nigerian users writing product reviews. "
            "For each row in the 'Evaluation' sheet you see one persona, one product, and two candidate reviews "
            "(Review A and Review B). Your job is to decide which review feels more like an AUTHENTIC Nigerian "
            "user matching that specific persona.",
            False, 11,
        ),
        ("", False, 11),
        ("How to vote", True, 14),
        ("1. Open the 'Evaluation' sheet (tab at the bottom of this window).", False, 11),
        ("2. Read the persona context (register tier, location, occupation).", False, 11),
        ("3. Read the product details.", False, 11),
        ("4. Read Review A and Review B carefully.", False, 11),
        (
            "5. In the 'Your vote' column, pick from the dropdown:",
            False, 11,
        ),
        ("       A_better  — Review A is more authentic", False, 11),
        ("       B_better  — Review B is more authentic", False, 11),
        ("       Equal     — Both equally authentic (or equally bad)", False, 11),
        ("       Skip      — You can't judge this one (e.g. content unclear)", False, 11),
        (
            "6. (Optional) Add a short reason in the 'Comments' column — even one phrase like "
            "'A used Pidgin correctly' or 'B made up a fake brand' is valuable.",
            False, 11,
        ),
        ("", False, 11),
        ("What 'authentic' means here", True, 14),
        (
            "Authentic = sounds like a real Nigerian person in this persona's register / occupation / region "
            "would actually write. Pay attention to:", False, 11,
        ),
        ("   • register tier (Pidgin vs code-mixed vs Nigerian English vs standard English)", False, 11),
        ("   • intensity calibration (does 'wahala' mean what it should?)", False, 11),
        ("   • cultural specifics (Owambe references? Alhamdulillah? Nollywood lingo?)", False, 11),
        ("   • whether the persona's stated priorities (value, durability, etc.) actually show up", False, 11),
        ("   • whether the rating chosen makes sense given the review's content", False, 11),
        ("", False, 11),
        ("Time estimate", True, 14),
        (
            "Each row takes ~45-60 seconds. 50 rows ≈ 35-45 minutes total. You don't need to do it in one sitting "
            "— save and reopen the file as needed.", False, 11,
        ),
        ("", False, 11),
        ("Blinding", True, 14),
        (
            "You do NOT know which review came from which model. The labels are randomised. Please don't try to "
            "guess — just rate authenticity. Your votes are pooled with other raters' anonymously.",
            False, 11,
        ),
        ("", False, 11),
        ("Return the filled file", True, 14),
        (
            "Save your filled file as `human_eval_<yourname>.xlsx` (e.g. `human_eval_ashinze.xlsx`) and send back "
            "to the team. We do not need any other formatting — just the Evaluation sheet with the 'Your vote' "
            "column filled in.", False, 11,
        ),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if size >= 14:
            ws.row_dimensions[i].height = 24
        elif text:
            ws.row_dimensions[i].height = max(15, int(20 + len(text) / 80 * 14))


def _write_personas_sheet(wb, personas: list[dict]) -> None:
    ws = wb.create_sheet("Personas")
    headers = ["persona_id", "register_tier", "location", "age", "occupation",
               "primary aspects (top 3)", "register markers"]
    widths = [22, 18, 25, 8, 35, 35, 40]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for i, p in enumerate(sorted(personas, key=lambda x: x.get("user_id", "")), start=2):
        d = p.get("demographics", {}) or {}
        ap = p.get("aspect_priority", {}) or {}
        top_aspects = ", ".join(a for a, _ in sorted(ap.items(), key=lambda x: -x[1])[:3])
        markers = ", ".join(p.get("register_markers", [])[:5])
        ws.cell(row=i, column=1, value=p.get("user_id"))
        ws.cell(row=i, column=2, value=p.get("register_tier"))
        ws.cell(row=i, column=3, value=d.get("location", ""))
        ws.cell(row=i, column=4, value=d.get("age_range", ""))
        ws.cell(row=i, column=5, value=d.get("occupation", ""))
        ws.cell(row=i, column=6, value=top_aspects)
        ws.cell(row=i, column=7, value=markers)
        for col in range(1, 8):
            ws.cell(row=i, column=col).alignment = WRAP
            ws.cell(row=i, column=col).border = BORDER
        ws.row_dimensions[i].height = 30
    ws.freeze_panes = "A2"


def _write_evaluation_sheet(wb, pairs: list[dict]) -> None:
    ws = wb.create_sheet("Evaluation")
    headers = [
        "pair_id", "persona_id", "register_tier", "location",
        "product_title", "product_category", "product_price (NGN)",
        "Review A", "Rating A",
        "Review B", "Rating B",
        "Your vote", "Comments (optional)",
    ]
    widths = [10, 22, 16, 22, 30, 20, 14, 60, 8, 60, 8, 14, 30]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # Data validation for 'Your vote' (column 12)
    dv = DataValidation(
        type="list",
        formula1='"A_better,B_better,Equal,Skip"',
        showDropDown=False,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Pick from dropdown",
        error="Use A_better / B_better / Equal / Skip",
    )
    ws.add_data_validation(dv)

    for i, p in enumerate(pairs, start=2):
        loc = (p.get("demographics") or {}).get("location", "")
        price = p.get("product_price")
        price_str = f"₦{int(price):,}" if price else "—"
        row = [
            p["pair_id"],
            p["persona_id"],
            p["register_tier"],
            loc,
            p["product_title"],
            p["product_category"],
            price_str,
            p["left"]["review"],
            p["left"]["rating"],
            p["right"]["review"],
            p["right"]["rating"],
            "",  # Your vote (empty, dropdown)
            "",  # Comments
        ]
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = WRAP
            c.border = BORDER
        # Add data validation to "Your vote" cell
        dv.add(ws.cell(row=i, column=12))
        # Tall rows for reading reviews
        ws.row_dimensions[i].height = 160

    ws.freeze_panes = "H2"  # freeze headers and the persona/product context columns


def write_workbook(pairs: list[dict], personas: list[dict], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_instructions_sheet(wb.active)
    _write_personas_sheet(wb, personas)
    _write_evaluation_sheet(wb, pairs)
    wb.save(out)


def write_master_json(pairs: list[dict], out: Path) -> None:
    """Master file — has the model→side mapping. NOT shareable."""
    master = [
        {
            "pair_id": p["pair_id"],
            "left_model": p["left"]["_model"],
            "right_model": p["right"]["_model"],
            "persona_id": p["persona_id"],
            "register_tier": p["register_tier"],
            "product_title": p["product_title"],
        }
        for p in pairs
    ]
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(master, indent=2, ensure_ascii=False))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--n", type=int, default=50,
                    help="Number of A/B pairs (default 50)")
    ap.add_argument("--model-a", type=str, default=MODEL_A_DEFAULT,
                    help="Model A spec (default: NaijaReviewer-8B via LM Studio)")
    ap.add_argument("--model-b", type=str, default=MODEL_B_DEFAULT,
                    help="Model B spec (default: Claude Sonnet 4)")
    ap.add_argument("--concurrency", type=int, default=2)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    pairs = asyncio.run(build_all_pairs(args.n, args.model_a, args.model_b,
                                          args.concurrency, args.seed))
    if not pairs:
        logger.error("no pairs built")
        return 2

    personas = _load_personas()
    write_workbook(pairs, personas, OUT_XLSX)
    write_master_json(pairs, OUT_MASTER)

    logger.info("💾 shareable workbook → %s", OUT_XLSX)
    logger.info("💾 master answer key  → %s (KEEP LOCAL)", OUT_MASTER)
    logger.info("")
    logger.info("Share %s with raters. They fill the 'Your vote' column.", OUT_XLSX.name)
    logger.info("Returned files go in paper/human_eval_returned/. Aggregate with:")
    logger.info("  python scripts/aggregate_human_eval_xlsx.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
