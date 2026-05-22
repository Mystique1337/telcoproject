"""Generate a shareable Excel workbook for the Task B (recommendation)
contextual-relevance human eval: NaijaReviewer-8B vs Claude Sonnet 4
(or any two re-rankers).

This is the recommendation-side analogue of build_human_eval_xlsx.py. Instead
of comparing two reviews, raters compare two ranked product lists generated for
the same Nigerian persona and judge which list is more contextually relevant
for that persona, plus rate each list's relevance on a 1-5 scale.

Output:
  paper/task_b_human_eval_template.xlsx   - shareable, NO model labels visible
  paper/task_b_human_eval_master.json     - local-only, has model->side mapping

Workbook structure (3 sheets):
  1. Instructions - what to do, how to vote, time estimate.
  2. Personas - quick reference table of the 24 persona archetypes.
  3. Evaluation - N rows, each with persona context, recommendation List A,
     List B, a 'Which list' dropdown, and a 1-5 relevance score per list.

The two lists are randomly side-swapped per row (50/50 chance List A is the
fine-tune or Claude) so raters cannot pattern-match.

Requires the FastAPI service running locally (make serve) so /recommend is
reachable, and a populated Pinecone index (or local Chroma fallback).

Run:
  # Default: one scenario per persona (24), NaijaReviewer-8B vs Claude
  python scripts/build_task_b_human_eval_xlsx.py

  # More scenarios for stronger statistics
  python scripts/build_task_b_human_eval_xlsx.py --n 48

Then share paper/task_b_human_eval_template.xlsx with raters; each fills the
'Which list' column and the two relevance columns. Renamed copies (e.g.
task_b_human_eval_ashinze.xlsx) come back to paper/task_b_human_eval_returned/,
and aggregate_task_b_human_eval_xlsx.py tallies them.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import sys
from pathlib import Path
from typing import Any

import httpx

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass

import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("build_task_b_xlsx")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s - %(message)s")

API_URL = "http://localhost:8765"
PERSONAS_DIR = PROJECT_ROOT / "data" / "sample" / "personas"

OUT_XLSX = PROJECT_ROOT / "paper" / "task_b_human_eval_template.xlsx"
OUT_MASTER = PROJECT_ROOT / "paper" / "task_b_human_eval_master.json"

MODEL_A_DEFAULT = "modal:naija-reviewer-8b"
MODEL_B_DEFAULT = "anthropic:claude-sonnet-4-20250514"

DEFAULT_DOMAIN = "jumia"
TOP_K = 5


# --------------------------------------------------------------------------- #
# Data loading                                                                 #
# --------------------------------------------------------------------------- #

def _load_personas() -> list[dict]:
    out = []
    for p in sorted(PERSONAS_DIR.glob("*.json")):
        try:
            out.append(json.loads(p.read_text(encoding="utf-8")))
        except Exception:
            continue
    return out


# --------------------------------------------------------------------------- #
# API calls (both re-rankers in parallel per scenario)                          #
# --------------------------------------------------------------------------- #

async def _recommend(client: httpx.AsyncClient, persona: dict, domain: str,
                     k: int, reranker: str) -> dict[str, Any]:
    try:
        r = await client.post(
            f"{API_URL}/recommend",
            json={"persona": persona, "domain": domain, "k": k,
                  "reranker_override": reranker},
            timeout=240,
        )
        r.raise_for_status()
        return r.json()
    except Exception as e:  # noqa: BLE001
        return {"error": str(e)[:200]}


def _format_list(resp: dict) -> str:
    """Render a recommendation response as a numbered, human-readable list."""
    items = resp.get("recommendations") or []
    lines = []
    for it in items:
        title = (it.get("title") or it.get("product_id") or "?")[:90]
        price = it.get("price_naira")
        price_str = f"  (NGN {int(price):,})" if price else ""
        lines.append(f"{it.get('rank', len(lines) + 1)}. {title}{price_str}")
    return "\n".join(lines) if lines else "(no recommendations returned)"


async def _build_scenario(client: httpx.AsyncClient, sem: asyncio.Semaphore,
                          persona: dict, domain: str, k: int,
                          model_a: str, model_b: str) -> dict | None:
    async with sem:
        a_task = _recommend(client, persona, domain, k, model_a)
        b_task = _recommend(client, persona, domain, k, model_b)
        a, b = await asyncio.gather(a_task, b_task)
    if "error" in a or "error" in b or not a.get("recommendations") or not b.get("recommendations"):
        logger.warning("  scenario failed: %s -> A=%s B=%s",
                       persona.get("user_id", "?"),
                       "ok" if a.get("recommendations") else "err",
                       "ok" if b.get("recommendations") else "err")
        return None
    return {"persona": persona, "domain": domain,
            "model_a_resp": a, "model_b_resp": b}


async def build_all_scenarios(n: int, model_a: str, model_b: str, domain: str,
                              k: int, concurrency: int, seed: int) -> list[dict]:
    rng = random.Random(seed)
    personas = _load_personas()
    if not personas:
        logger.error("no personas under %s", PERSONAS_DIR)
        return []

    # Stratified: cycle through personas so each gets ~equal coverage.
    persona_cycle = list(personas)
    rng.shuffle(persona_cycle)
    chosen = [persona_cycle[i % len(persona_cycle)] for i in range(n)]

    sem = asyncio.Semaphore(concurrency)
    async with httpx.AsyncClient() as client:
        results = await asyncio.gather(*[
            _build_scenario(client, sem, p, domain, k, model_a, model_b)
            for p in chosen
        ])
    results = [r for r in results if r]

    rows: list[dict] = []
    for i, r in enumerate(results, start=1):
        swap = rng.random() < 0.5
        left_model, right_model = (model_b, model_a) if swap else (model_a, model_b)
        left_resp = r["model_b_resp"] if swap else r["model_a_resp"]
        right_resp = r["model_a_resp"] if swap else r["model_b_resp"]
        persona = r["persona"]
        rows.append({
            "scenario_id": f"S{i:03d}",
            "persona_id": persona.get("user_id"),
            "register_tier": persona.get("register_tier"),
            "demographics": persona.get("demographics", {}),
            "domain": r["domain"],
            "left": {"_model": left_model, "list": _format_list(left_resp)},
            "right": {"_model": right_model, "list": _format_list(right_resp)},
        })
    logger.info("built %d/%d scenarios", len(rows), n)
    return rows


# --------------------------------------------------------------------------- #
# Workbook styling                                                              #
# --------------------------------------------------------------------------- #

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Task B - Recommendation Relevance Evaluation", True, 14),
        ("", False, 11),
        ("We are evaluating two AI recommendation engines for Nigerian shoppers. "
         "For each row in the 'Evaluation' sheet you see one persona and two ranked "
         "product lists (List A and List B) generated for that same persona. Your job "
         "is to judge which list is more contextually relevant for that specific "
         "Nigerian shopper.", False, 11),
        ("", False, 11),
        ("What to do", True, 12),
        ("1. Open the 'Evaluation' sheet (tab at the bottom).", False, 11),
        ("2. Read the persona context (register tier, location, age, occupation) and "
         "check the 'Personas' sheet for that persona's stated priorities.", False, 11),
        ("3. Read List A and List B (each is a ranked set of products with prices).", False, 11),
        ("4. In 'Which list', pick from the dropdown which list better fits this persona:", False, 11),
        ("       A_better / B_better / Equal / Skip", False, 11),
        ("5. In 'Relevance A' and 'Relevance B', rate each list 1-5 for how well it "
         "suits this persona (1 = irrelevant, 5 = highly relevant).", False, 11),
        ("6. (Optional) Add a short reason in 'Comments'.", False, 11),
        ("", False, 11),
        ("What 'contextually relevant' means", True, 12),
        ("   - Do the products match the persona's stated priorities (value, durability, "
         "quality, etc.)?", False, 11),
        ("   - Are the prices appropriate for this persona's likely budget?", False, 11),
        ("   - Would a real Nigerian shopper of this profile plausibly want these items?", False, 11),
        ("", False, 11),
        ("You do NOT know which engine produced which list. The labels are randomised. "
         "Please do not try to guess. Save the file with your name (e.g. "
         "task_b_human_eval_yourname.xlsx) and send it back. We only need the "
         "Evaluation sheet filled in.", False, 11),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = WRAP
        cell.font = Font(bold=bold, size=size)
        if bold and size >= 14:
            ws.row_dimensions[i].height = 24
        else:
            ws.row_dimensions[i].height = max(15, int(20 + len(text) / 80 * 14))


def _write_personas_sheet(wb, personas: list[dict]) -> None:
    ws = wb.create_sheet("Personas")
    headers = ["persona_id", "register_tier", "location", "age", "occupation",
               "top_aspects", "markers"]
    widths = [22, 18, 18, 12, 22, 28, 28]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for i, p in enumerate(sorted(personas, key=lambda x: x.get("user_id", "")), start=2):
        d = p.get("demographics", {}) or {}
        aspects = p.get("aspect_priority", {}) or {}
        top_aspects = ", ".join(
            f"{k}:{v:.2f}" for k, v in sorted(aspects.items(), key=lambda kv: -kv[1])[:3]
        )
        markers = ", ".join((p.get("register_markers") or [])[:6])
        ws.cell(row=i, column=1, value=p.get("user_id"))
        ws.cell(row=i, column=2, value=p.get("register_tier"))
        ws.cell(row=i, column=3, value=d.get("location", ""))
        ws.cell(row=i, column=4, value=d.get("age_range", ""))
        ws.cell(row=i, column=5, value=d.get("occupation", ""))
        ws.cell(row=i, column=6, value=top_aspects)
        ws.cell(row=i, column=7, value=markers)
        for col in range(1, 8):
            ws.cell(row=i, column=col).alignment = WRAP
            ws.cell(row=i, column=col).border = BORDER
        ws.row_dimensions[i].height = 30
    ws.freeze_panes = "A2"


def _write_evaluation_sheet(wb, rows: list[dict]) -> None:
    ws = wb.create_sheet("Evaluation")
    headers = [
        "scenario_id", "persona_id", "register_tier", "location", "occupation",
        "List A", "List B",
        "Which list", "Relevance A (1-5)", "Relevance B (1-5)", "Comments (optional)",
    ]
    widths = [12, 22, 16, 18, 22, 46, 46, 14, 16, 16, 30]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    # 'Which list' dropdown (column 8)
    dv_vote = DataValidation(
        type="list", formula1='"A_better,B_better,Equal,Skip"',
        showDropDown=False, allow_blank=True, showErrorMessage=True,
        errorTitle="Pick from dropdown", error="Use A_better / B_better / Equal / Skip",
    )
    ws.add_data_validation(dv_vote)
    # Relevance dropdowns (columns 9, 10)
    dv_rel = DataValidation(
        type="list", formula1='"1,2,3,4,5"',
        showDropDown=False, allow_blank=True, showErrorMessage=True,
        errorTitle="Pick 1-5", error="Rate relevance from 1 (low) to 5 (high)",
    )
    ws.add_data_validation(dv_rel)

    for i, r in enumerate(rows, start=2):
        loc = (r.get("demographics") or {}).get("location", "")
        occ = (r.get("demographics") or {}).get("occupation", "")
        row = [
            r["scenario_id"], r["persona_id"], r["register_tier"], loc, occ,
            r["left"]["list"], r["right"]["list"],
            "", "", "", "",
        ]
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = WRAP
            c.border = BORDER
        dv_vote.add(ws.cell(row=i, column=8))
        dv_rel.add(ws.cell(row=i, column=9))
        dv_rel.add(ws.cell(row=i, column=10))
        ws.row_dimensions[i].height = 150

    ws.freeze_panes = "F2"


def write_workbook(rows: list[dict], personas: list[dict], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_instructions_sheet(wb.active)
    _write_personas_sheet(wb, personas)
    _write_evaluation_sheet(wb, rows)
    wb.save(out)


def write_master_json(rows: list[dict], out: Path) -> None:
    """Master file - has the model->side mapping. NOT shareable."""
    master = [
        {
            "scenario_id": r["scenario_id"],
            "left_model": r["left"]["_model"],
            "right_model": r["right"]["_model"],
            "persona_id": r["persona_id"],
            "register_tier": r["register_tier"],
            "domain": r["domain"],
        }
        for r in rows
    ]
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(master, indent=2, ensure_ascii=False))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--n", type=int, default=24,
                    help="Number of scenarios (default 24: one per persona)")
    ap.add_argument("--model-a", type=str, default=MODEL_A_DEFAULT,
                    help="Re-ranker A spec (default: NaijaReviewer-8B on Modal)")
    ap.add_argument("--model-b", type=str, default=MODEL_B_DEFAULT,
                    help="Re-ranker B spec (default: Claude Sonnet 4)")
    ap.add_argument("--domain", type=str, default=DEFAULT_DOMAIN)
    ap.add_argument("--k", type=int, default=TOP_K)
    ap.add_argument("--concurrency", type=int, default=2)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    rows = asyncio.run(build_all_scenarios(
        args.n, args.model_a, args.model_b, args.domain, args.k,
        args.concurrency, args.seed))
    if not rows:
        logger.error("no scenarios built (is `make serve` running?)")
        return 2

    personas = _load_personas()
    write_workbook(rows, personas, OUT_XLSX)
    write_master_json(rows, OUT_MASTER)

    logger.info("shareable workbook -> %s", OUT_XLSX)
    logger.info("master answer key  -> %s (KEEP LOCAL)", OUT_MASTER)
    logger.info("")
    logger.info("Share %s with raters. They fill 'Which list' + relevance columns.",
                OUT_XLSX.name)
    logger.info("Returned files go in paper/task_b_human_eval_returned/. Aggregate with:")
    logger.info("  python scripts/aggregate_task_b_human_eval_xlsx.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
