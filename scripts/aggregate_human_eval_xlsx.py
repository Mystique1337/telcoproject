"""Aggregate filled human-eval XLSX files into a single summary.

Reads back files from ``paper/human_eval_returned/`` (or a directory passed
on the CLI) — each should be a copy of ``human_eval_template.xlsx`` with the
'Your vote' column filled in (A_better / B_better / Equal / Skip).

Cross-references with ``paper/human_eval_pairs_master.json`` (the master
answer key) to figure out which model each vote represents.

Output:
  paper/human_eval_summary.json  — per-rater + overall win-rate, Wilson 95% CI,
                                    Krippendorff α for inter-rater agreement
  paper/human_eval_summary.md    — paper-ready markdown table

Run:
  python scripts/aggregate_human_eval_xlsx.py
  python scripts/aggregate_human_eval_xlsx.py --returned-dir /path/to/files/
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logger = logging.getLogger("agg_xlsx")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s — %(message)s")

DEFAULT_RETURNED_DIR = PROJECT_ROOT / "paper" / "human_eval_returned"
MASTER_PATH = PROJECT_ROOT / "paper" / "human_eval_pairs_master.json"
OUT_JSON = PROJECT_ROOT / "paper" / "human_eval_summary.json"
OUT_MD = PROJECT_ROOT / "paper" / "human_eval_summary.md"

# Match what was set when the template was generated
MODEL_A_DEFAULT = "lmstudio:naija-reviewer-8b"


# --------------------------------------------------------------------------- #
# Stats helpers                                                                 #
# --------------------------------------------------------------------------- #

def _wilson_ci(wins: int, n: int, confidence: float = 0.95) -> tuple[float, float]:
    if n == 0:
        return (float("nan"), float("nan"))
    p = wins / n
    z = 1.96  # 95%
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n) / denom
    return (max(0.0, centre - half), min(1.0, centre + half))


def _krippendorff_alpha_nominal(rater_votes: dict[str, dict[str, str]]) -> float:
    """Nominal Krippendorff's α over pairs. Skips/blank votes ignored."""
    by_pair: dict[str, dict[str, str]] = {}
    for rater, votes in rater_votes.items():
        for pid, v in votes.items():
            if v in ("Skip", "SKIP", "", None):
                continue
            by_pair.setdefault(pid, {})[rater] = v
    categories = sorted({v for raters in by_pair.values() for v in raters.values()})
    cat_index = {c: i for i, c in enumerate(categories)}
    k = len(categories)
    if k < 2:
        return float("nan")
    coincidence = [[0.0] * k for _ in range(k)]
    for raters in by_pair.values():
        m = len(raters)
        if m < 2:
            continue
        vs = list(raters.values())
        for i in range(m):
            for j in range(m):
                if i == j:
                    continue
                coincidence[cat_index[vs[i]]][cat_index[vs[j]]] += 1.0 / (m - 1)
    n_c = [sum(coincidence[c]) for c in range(k)]
    n_total = sum(n_c)
    if n_total == 0:
        return float("nan")
    do = sum(coincidence[c1][c2] for c1 in range(k) for c2 in range(k) if c1 != c2) / n_total
    de = 1.0 - sum((nc / n_total) ** 2 for nc in n_c)
    if de == 0:
        return float("nan")
    return 1 - do / de


# --------------------------------------------------------------------------- #
# Workbook reader                                                              #
# --------------------------------------------------------------------------- #

def _normalize_vote(v) -> str:
    """Tolerant vote parser — accepts case-insensitive labels + common synonyms."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    if not s:
        return ""
    if s in ("a", "a_better", "a better", "review a", "a is better", "left"):
        return "A_better"
    if s in ("b", "b_better", "b better", "review b", "b is better", "right"):
        return "B_better"
    if s in ("equal", "=", "tie", "same", "both"):
        return "Equal"
    if s in ("skip", "s", "n/a", "na", "-", "x"):
        return "Skip"
    return ""


def _read_rater_xlsx(path: Path) -> dict[str, str]:
    """Read pair_id → vote from a filled template."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Evaluation" not in wb.sheetnames:
        logger.warning("  %s missing 'Evaluation' sheet — skipping", path.name)
        return {}
    ws = wb["Evaluation"]
    headers = [c.value for c in ws[1]]
    try:
        pid_col = headers.index("pair_id") + 1
        vote_col = headers.index("Your vote") + 1
    except ValueError:
        logger.warning("  %s headers don't match template — skipping", path.name)
        return {}
    votes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col - 1]
        v = _normalize_vote(row[vote_col - 1])
        if pid and v:
            votes[str(pid)] = v
    return votes


def _extract_rater_name(filename: str) -> str:
    """human_eval_ashinze.xlsx → 'ashinze'. Fallback: filename stem."""
    stem = Path(filename).stem
    m = re.match(r"human[_-]eval[_-](.+)", stem, re.IGNORECASE)
    if m:
        return m.group(1)
    return stem


# --------------------------------------------------------------------------- #
# Main                                                                          #
# --------------------------------------------------------------------------- #

def aggregate(returned_dir: Path) -> int:
    if not MASTER_PATH.exists():
        logger.error("master file missing: %s", MASTER_PATH)
        logger.error("Run scripts/build_human_eval_xlsx.py first to generate it.")
        return 2
    master_records = json.loads(MASTER_PATH.read_text())
    side_by_pair = {m["pair_id"]: m for m in master_records}
    n_pairs = len(side_by_pair)

    if not returned_dir.exists():
        logger.error("returned dir does not exist: %s", returned_dir)
        logger.error("Create it and drop filled xlsx files there.")
        return 2

    xlsx_files = sorted(returned_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning("no .xlsx files in %s", returned_dir)
        logger.warning("Drop filled human-eval files there and re-run.")
        return 1

    logger.info("reading %d returned xlsx files...", len(xlsx_files))
    rater_votes: dict[str, dict[str, str]] = {}
    for f in xlsx_files:
        rater = _extract_rater_name(f.name)
        votes = _read_rater_xlsx(f)
        rater_votes[rater] = votes
        logger.info("  %s: %d votes", rater, len(votes))

    # Tally per-rater
    per_rater: dict[str, dict] = {}
    overall_naija = 0
    overall_claude = 0
    overall_ties = 0
    overall_decisive = 0
    for rater, votes in rater_votes.items():
        n = c = t = decisive = 0
        for pid, v in votes.items():
            if v == "Skip" or v == "":
                continue
            if v == "Equal":
                t += 1
                continue
            decisive += 1
            sb = side_by_pair.get(pid)
            if not sb:
                continue
            chosen_side = "left_model" if v == "A_better" else "right_model"
            picked = sb[chosen_side]
            if picked == MODEL_A_DEFAULT:
                n += 1
            else:
                c += 1
        overall_naija += n
        overall_claude += c
        overall_ties += t
        overall_decisive += decisive
        wlo, whi = _wilson_ci(n, decisive)
        per_rater[rater] = {
            "naija_wins": n, "claude_wins": c, "ties": t,
            "decisive": decisive,
            "naija_win_rate": (n / max(decisive, 1)) if decisive else None,
            "naija_win_rate_ci95": [wlo, whi],
        }

    overall_win_rate = (overall_naija / max(overall_decisive, 1)) if overall_decisive else None
    wlo, whi = _wilson_ci(overall_naija, overall_decisive)
    alpha = _krippendorff_alpha_nominal(rater_votes)

    summary = {
        "n_raters": len(rater_votes),
        "raters": list(rater_votes.keys()),
        "n_pairs_in_master": n_pairs,
        "per_rater": per_rater,
        "overall": {
            "naija_wins": overall_naija,
            "claude_wins": overall_claude,
            "ties": overall_ties,
            "decisive_total": overall_decisive,
            "naija_win_rate": overall_win_rate,
            "naija_win_rate_ci95": [wlo, whi],
            "krippendorff_alpha_nominal": alpha,
        },
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(summary, indent=2, default=str))
    logger.info("💾 %s", OUT_JSON)

    # Markdown
    lines = [
        "# Human Eval Summary — NaijaReviewer-8B vs Claude Sonnet 4 (blind A/B)",
        "",
        f"- Raters: **{len(rater_votes)}** ({', '.join(rater_votes) or 'none'})",
        f"- Pairs in master: **{n_pairs}**",
        f"- Decisive votes (excluding Equal/Skip): **{overall_decisive}**",
        f"- Ties (Equal votes): **{overall_ties}**",
        "",
        "## Headline",
        "",
        (f"**NaijaReviewer-8B win-rate: {overall_win_rate:.1%}  "
         f"(95% CI [{wlo:.1%}, {whi:.1%}])**"
         if overall_win_rate is not None else
         "_No decisive votes recorded yet._"),
        "",
        f"Inter-rater agreement (Krippendorff α, nominal): **{alpha:.3f}**" if alpha == alpha else "Krippendorff α: n/a",
        "",
        "## Per-rater breakdown",
        "",
        "| Rater | NaijaReviewer wins | Claude wins | Ties | Decisive | NaijaReviewer win-rate | 95% CI |",
        "|---|---|---|---|---|---|---|",
    ]
    for rater, r in per_rater.items():
        wlo_r, whi_r = r["naija_win_rate_ci95"]
        wr = r["naija_win_rate"]
        lines.append(
            f"| {rater} | {r['naija_wins']} | {r['claude_wins']} | {r['ties']} | "
            f"{r['decisive']} | "
            + (f"{wr:.1%}" if wr is not None else "—")
            + f" | [{wlo_r:.1%}, {whi_r:.1%}] |"
        )
    OUT_MD.write_text("\n".join(lines))
    logger.info("💾 %s", OUT_MD)

    print("\n" + "=" * 70)
    print(OUT_MD.read_text())
    print("=" * 70)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--returned-dir", type=Path, default=DEFAULT_RETURNED_DIR,
                    help="Directory with returned filled xlsx files")
    args = ap.parse_args()
    return aggregate(args.returned_dir)


if __name__ == "__main__":
    raise SystemExit(main())
