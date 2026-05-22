"""Aggregate filled Task B (recommendation) relevance human-eval workbooks
into a single summary.

Reads back files from paper/task_b_human_eval_returned/ (or a directory passed
on the CLI) - each a copy of task_b_human_eval_template.xlsx with the
'Which list' column and the two relevance columns filled in.

Cross-references paper/task_b_human_eval_master.json (the answer key) to map
each List-A/List-B vote and relevance score back to the model that produced it.

Output:
  paper/task_b_human_eval_summary.json  - per-rater + overall win-rate, Wilson
                                          95% CI, mean relevance per model,
                                          Krippendorff alpha
  paper/task_b_human_eval_summary.md    - paper-ready markdown

Run:
  python scripts/aggregate_task_b_human_eval_xlsx.py
  python scripts/aggregate_task_b_human_eval_xlsx.py --returned-dir /path/to/files/
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logger = logging.getLogger("agg_task_b")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s - %(message)s")

DEFAULT_RETURNED_DIR = PROJECT_ROOT / "paper" / "task_b_human_eval_returned"
MASTER_PATH = PROJECT_ROOT / "paper" / "task_b_human_eval_master.json"
OUT_JSON = PROJECT_ROOT / "paper" / "task_b_human_eval_summary.json"
OUT_MD = PROJECT_ROOT / "paper" / "task_b_human_eval_summary.md"

# The fine-tune spec we want to report a win-rate FOR. Matches the builder's
# --model-a default; any spec containing "naija" is treated as the fine-tune.
FINE_TUNE_HINT = "naija"


# --------------------------------------------------------------------------- #
# Stats helpers                                                                 #
# --------------------------------------------------------------------------- #

def _wilson_ci(wins: int, n: int) -> tuple[float, float]:
    if n == 0:
        return (float("nan"), float("nan"))
    p = wins / n
    z = 1.96
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n) / denom
    return (max(0.0, centre - half), min(1.0, centre + half))


def _krippendorff_alpha_nominal(rater_votes: dict[str, dict[str, str]]) -> float:
    """Nominal Krippendorff's alpha over scenarios. Skip/blank ignored."""
    by_item: dict[str, dict[str, str]] = {}
    for rater, votes in rater_votes.items():
        for sid, v in votes.items():
            if v in ("Skip", "SKIP", "", None):
                continue
            by_item.setdefault(sid, {})[rater] = v
    categories = sorted({v for raters in by_item.values() for v in raters.values()})
    cat_index = {c: i for i, c in enumerate(categories)}
    k = len(categories)
    if k < 2:
        return float("nan")
    coincidence = [[0.0] * k for _ in range(k)]
    for raters in by_item.values():
        m = len(raters)
        if m < 2:
            continue
        vs = list(raters.values())
        for i in range(m):
            for j in range(m):
                if i == j:
                    continue
                coincidence[cat_index[vs[i]]][cat_index[vs[j]]] += 1.0 / (m - 1)
    n_c = [sum(coincidence[c]) for c in range(k)]
    n_total = sum(n_c)
    if n_total == 0:
        return float("nan")
    do = sum(coincidence[c1][c2] for c1 in range(k) for c2 in range(k) if c1 != c2) / n_total
    de = 1.0 - sum((nc / n_total) ** 2 for nc in n_c)
    if de == 0:
        return float("nan")
    return 1 - do / de


# --------------------------------------------------------------------------- #
# Workbook reader                                                              #
# --------------------------------------------------------------------------- #

def _normalize_vote(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    if not s:
        return ""
    if s in ("a", "a_better", "a better", "list a", "a is better", "left"):
        return "A_better"
    if s in ("b", "b_better", "b better", "list b", "b is better", "right"):
        return "B_better"
    if s in ("equal", "=", "tie", "same", "both"):
        return "Equal"
    if s in ("skip", "s", "n/a", "na", "-", "x"):
        return "Skip"
    return ""


def _parse_rel(v) -> int | None:
    try:
        n = int(round(float(str(v).strip())))
    except (TypeError, ValueError):
        return None
    return n if 1 <= n <= 5 else None


def _read_rater_xlsx(path: Path) -> dict[str, dict]:
    """Read scenario_id -> {vote, rel_a, rel_b} from a filled template."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Evaluation" not in wb.sheetnames:
        logger.warning("  %s missing 'Evaluation' sheet - skipping", path.name)
        return {}
    ws = wb["Evaluation"]
    headers = [c.value for c in ws[1]]
    try:
        sid_col = headers.index("scenario_id")
        vote_col = headers.index("Which list")
        rel_a_col = headers.index("Relevance A (1-5)")
        rel_b_col = headers.index("Relevance B (1-5)")
    except ValueError:
        logger.warning("  %s headers don't match template - skipping", path.name)
        return {}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = row[sid_col]
        if not sid:
            continue
        vote = _normalize_vote(row[vote_col])
        rel_a = _parse_rel(row[rel_a_col])
        rel_b = _parse_rel(row[rel_b_col])
        if vote or rel_a is not None or rel_b is not None:
            out[str(sid)] = {"vote": vote, "rel_a": rel_a, "rel_b": rel_b}
    return out


def _extract_rater_name(filename: str) -> str:
    stem = Path(filename).stem
    m = re.match(r"task[_-]?b[_-]human[_-]eval[_-](.+)", stem, re.IGNORECASE)
    if m:
        return m.group(1)
    return stem


# --------------------------------------------------------------------------- #
# Main                                                                          #
# --------------------------------------------------------------------------- #

def aggregate(returned_dir: Path) -> int:
    if not MASTER_PATH.exists():
        logger.error("master file missing: %s", MASTER_PATH)
        logger.error("Run scripts/build_task_b_human_eval_xlsx.py first.")
        return 2
    master = {m["scenario_id"]: m for m in json.loads(MASTER_PATH.read_text())}

    if not returned_dir.exists():
        logger.error("returned dir does not exist: %s", returned_dir)
        return 2
    files = sorted(p for p in returned_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not files:
        logger.error("no .xlsx files in %s", returned_dir)
        return 2

    # rater -> scenario_id -> which-model-vote ("naija"/"claude"/"Equal"/"Skip")
    rater_model_votes: dict[str, dict[str, str]] = {}
    # rater -> list of (fine_tune_rel, other_rel) per scenario
    rater_rel: dict[str, list[tuple[int, int]]] = defaultdict(list)
    per_rater = {}

    for f in files:
        rater = _extract_rater_name(f.name)
        rows = _read_rater_xlsx(f)
        if not rows:
            continue
        model_votes: dict[str, str] = {}
        ft_wins = other_wins = ties = decisive = 0
        ft_rels: list[int] = []
        other_rels: list[int] = []
        for sid, rec in rows.items():
            m = master.get(sid)
            if not m:
                continue
            left_is_ft = FINE_TUNE_HINT in (m["left_model"] or "").lower()
            # which side is the fine-tune
            ft_side = "A" if left_is_ft else "B"
            # map relevance scores to models
            rel_a, rel_b = rec.get("rel_a"), rec.get("rel_b")
            ft_rel = rel_a if ft_side == "A" else rel_b
            other_rel = rel_b if ft_side == "A" else rel_a
            if ft_rel is not None:
                ft_rels.append(ft_rel)
            if other_rel is not None:
                other_rels.append(other_rel)
            if ft_rel is not None and other_rel is not None:
                rater_rel[rater].append((ft_rel, other_rel))
            # map the which-list vote to a model
            vote = rec.get("vote") or ""
            if vote == "Equal":
                model_votes[sid] = "Equal"; ties += 1
            elif vote == "Skip" or vote == "":
                if vote == "Skip":
                    model_votes[sid] = "Skip"
            elif vote == "A_better":
                won_ft = (ft_side == "A")
                model_votes[sid] = "naija" if won_ft else "other"
                decisive += 1; ft_wins += won_ft; other_wins += (not won_ft)
            elif vote == "B_better":
                won_ft = (ft_side == "B")
                model_votes[sid] = "naija" if won_ft else "other"
                decisive += 1; ft_wins += won_ft; other_wins += (not won_ft)

        rater_model_votes[rater] = model_votes
        wr = (ft_wins / decisive) if decisive else float("nan")
        lo, hi = _wilson_ci(ft_wins, decisive)
        per_rater[rater] = {
            "fine_tune_wins": ft_wins, "other_wins": other_wins, "ties": ties,
            "decisive": decisive, "win_rate": wr, "ci95": [lo, hi],
            "mean_rel_fine_tune": (sum(ft_rels) / len(ft_rels)) if ft_rels else float("nan"),
            "mean_rel_other": (sum(other_rels) / len(other_rels)) if other_rels else float("nan"),
            "n_rel_fine_tune": len(ft_rels), "n_rel_other": len(other_rels),
        }
        logger.info("  %-14s decisive=%d  win-rate=%.1f%%  relFT=%.2f relOther=%.2f",
                    rater, decisive, 100 * wr if decisive else 0.0,
                    per_rater[rater]["mean_rel_fine_tune"],
                    per_rater[rater]["mean_rel_other"])

    # ---- pooled ----
    tot_ft = sum(r["fine_tune_wins"] for r in per_rater.values())
    tot_other = sum(r["other_wins"] for r in per_rater.values())
    tot_ties = sum(r["ties"] for r in per_rater.values())
    tot_dec = sum(r["decisive"] for r in per_rater.values())
    overall_wr = (tot_ft / tot_dec) if tot_dec else float("nan")
    lo, hi = _wilson_ci(tot_ft, tot_dec)

    all_ft_rel = [a for pairs in rater_rel.values() for (a, _) in pairs]
    all_other_rel = [b for pairs in rater_rel.values() for (_, b) in pairs]
    mean_ft = (sum(all_ft_rel) / len(all_ft_rel)) if all_ft_rel else float("nan")
    mean_other = (sum(all_other_rel) / len(all_other_rel)) if all_other_rel else float("nan")
    alpha = _krippendorff_alpha_nominal(rater_model_votes)

    summary = {
        "n_raters": len(per_rater),
        "n_scenarios_master": len(master),
        "fine_tune_hint": FINE_TUNE_HINT,
        "overall": {
            "fine_tune_wins": tot_ft, "other_wins": tot_other, "ties": tot_ties,
            "decisive": tot_dec, "win_rate": overall_wr, "ci95": [lo, hi],
            "mean_relevance_fine_tune": mean_ft, "mean_relevance_other": mean_other,
            "n_paired_relevance": len(all_ft_rel),
            "krippendorff_alpha": alpha,
        },
        "per_rater": per_rater,
    }
    OUT_JSON.write_text(json.dumps(summary, indent=2))
    _write_md(summary)
    logger.info("")
    logger.info("summary JSON -> %s", OUT_JSON)
    logger.info("summary MD   -> %s", OUT_MD)
    if tot_dec:
        logger.info("OVERALL fine-tune win-rate: %.1f%% (95%% CI [%.1f%%, %.1f%%])",
                    100 * overall_wr, 100 * lo, 100 * hi)
        logger.info("mean relevance: fine-tune %.2f vs other %.2f", mean_ft, mean_other)
    return 0


def _fmt_pct(x: float) -> str:
    return "n/a" if x != x else f"{100 * x:.1f}%"


def _fmt_num(x: float) -> str:
    return "n/a" if x != x else f"{x:.2f}"


def _write_md(s: dict) -> None:
    o = s["overall"]
    lines = [
        "# Task B Human Eval - Recommendation Relevance (blind A/B)",
        "",
        "NaijaReviewer-8B vs Claude Sonnet 4 as the recommendation re-ranker. "
        "Raters judged which of two blind, randomised product lists is more "
        "contextually relevant for each Nigerian persona, and rated each list 1-5.",
        "",
        f"- Raters: **{s['n_raters']}**",
        f"- Scenarios in master: **{s['n_scenarios_master']}**",
        f"- Decisive votes (excluding Equal/Skip): **{o['decisive']}**",
        f"- Ties (Equal votes): **{o['ties']}**",
        "",
        "## Headline",
        "",
        f"**NaijaReviewer-8B win-rate: {_fmt_pct(o['win_rate'])}  "
        f"(95% CI [{_fmt_pct(o['ci95'][0])}, {_fmt_pct(o['ci95'][1])}])**",
        "",
        f"Mean relevance (1-5): **NaijaReviewer-8B {_fmt_num(o['mean_relevance_fine_tune'])}** "
        f"vs **Claude {_fmt_num(o['mean_relevance_other'])}** "
        f"(over {o['n_paired_relevance']} paired ratings).",
        "",
        f"Inter-rater agreement (Krippendorff alpha, nominal): **{_fmt_num(o['krippendorff_alpha'])}**",
        "",
        "## Per-rater breakdown",
        "",
        "| Rater | Naija wins | Other wins | Ties | Decisive | Naija win-rate | 95% CI | rel Naija | rel Other |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for rater, r in sorted(s["per_rater"].items()):
        ci = f"[{_fmt_pct(r['ci95'][0])}, {_fmt_pct(r['ci95'][1])}]"
        lines.append(
            f"| {rater} | {r['fine_tune_wins']} | {r['other_wins']} | {r['ties']} | "
            f"{r['decisive']} | {_fmt_pct(r['win_rate'])} | {ci} | "
            f"{_fmt_num(r['mean_rel_fine_tune'])} | {_fmt_num(r['mean_rel_other'])} |"
        )
    OUT_MD.write_text("\n".join(lines) + "\n")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--returned-dir", type=Path, default=DEFAULT_RETURNED_DIR)
    args = ap.parse_args()
    return aggregate(args.returned_dir)


if __name__ == "__main__":
    raise SystemExit(main())
